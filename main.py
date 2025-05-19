import sys
import yaml
import datetime
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font


def interpolate_value(start, peak, taper_end, current_cycle, total_cycles, peak_cycle_idx=None, taper_cycles=2):
    """
    Interpolates a value for a workout (duration or mileage) based on the cycle position.
    If peak_cycle_idx is provided, use it for the peak, otherwise peak is at total_cycles - taper_cycles.
    """
    if current_cycle == 0:
        return start
    if peak_cycle_idx is not None:
        # Ramp up to peak, then taper
        if current_cycle < peak_cycle_idx:
            return start + (peak - start) * (current_cycle / peak_cycle_idx)
        elif current_cycle == peak_cycle_idx:
            return peak
        else:
            # Taper from peak to taper_end
            taper_span = total_cycles - peak_cycle_idx - 1
            if taper_span <= 0:
                return taper_end
            return peak + (taper_end - peak) * ((current_cycle - peak_cycle_idx) / taper_span)
    else:
        if current_cycle >= total_cycles - taper_cycles:
            return taper_end + (peak - taper_end) * max(0, (total_cycles - current_cycle - 1) / taper_cycles)
        else:
            return start + (peak - start) * (current_cycle / (total_cycles - taper_cycles))

def main():
    if len(sys.argv) < 2:
        print("Usage: python main2.py <plan.yaml>")
        sys.exit(1)
    with open(sys.argv[1]) as f:
        plan = yaml.safe_load(f)

    race_date = datetime.datetime.strptime(plan["race_date"], "%Y-%m-%d")
    start_plan = datetime.datetime.strptime(plan["start_plan"], "%Y-%m-%d")
    # max_long_run = plan["max_long_run"]
    days_per_microcycle = plan["days_per_microcycle"]
    run_descriptions = plan["run_descriptions"]
    cycle_descriptions = plan["cycle_descriptions"]
    cycles = plan["cycles"]
    total_cycles = len(cycles)

    # Find the peak cycle index (if any)
    peak_cycle_idx = None
    for idx, cycle in enumerate(cycles):
        if cycle.get("peak", False):
            peak_cycle_idx = idx
            break

    wb = Workbook()
    ws = wb.active
    ws.title = "Training Plan"

    # Run Type Descriptions section
    ws["A1"] = "Run Type Descriptions:"
    ws["A1"].font = Font(bold=True)
    row = 2
    for abbr, desc in run_descriptions.items():
        ws.append([abbr, desc.get("description", "")])
        row += 1
    ws.append([])
    row += 1

    est_avg_pace = plan.get("est_avg_pace", 10)  # default 10 min/mile if not set
    # Place est_avg_pace in a cell above the column headers
    ws[f"A{row}"] = "est_avg_pace (min/mile):"
    ws[f"B{row}"] = est_avg_pace
    row += 1
    # Table headers (no Workout Type Descriptions, milage -> mileage)
    headers = [
        "Workout ID",
        "Microcycle ID",
        "Intra-cycle Index",
        "Date",
        "Weekday",
        "WT",
        "mileage",
        "duration (minutes)",
    ]
    ws.append(headers)
    for cell in ws[row]:
        cell.alignment = Alignment(wrap_text=True)
        cell.font = Font(bold=True)
    row += 1

    workout_id = 0
    microcycle_id = 1
    day_pointer = start_plan
    microcycle_summaries = []
    week_summaries = []
    week_data = []
    microcycle_data = []

    # Track the last 'L' row index and its date
    last_l_row = None
    last_l_date = None
    microcycle_row_ranges = []  # List of (start_row, end_row) for each microcycle
    current_microcycle_start_row = None

    for cycle_idx, cycle in enumerate(cycles):
        cycle_type = cycle["type"]
        microcycle = cycle_descriptions[cycle_type]["microcycle"]
        for intra_idx, wt in enumerate(microcycle):
            date = day_pointer + datetime.timedelta(days=workout_id)
            if date > race_date:
                break
            if date == race_date:
                wt = "RD"
            if wt == "L":
                last_l_row = ws.max_row + 1  # 1-based row index in Excel
                last_l_date = date
            # Use spreadsheet formula for weekday (relative reference)
            weekday_formula = '=TEXT(OFFSET(INDIRECT("RC",FALSE),0,-1), "dddd")'
            desc = run_descriptions[wt]
            duration = None
            mileage = None
            mileage_constrained = False
            peak = desc.get("peak", {})
            start = desc.get("start", {})
            taper_end = desc.get("taper_end", {})
            if "miles" in peak and "miles" in start and "miles" in taper_end:
                mileage = round(interpolate_value(start["miles"], peak["miles"], taper_end["miles"], cycle_idx, total_cycles, peak_cycle_idx), 2)
                mileage_constrained = True
            elif "duration" in peak and "duration" in start and "duration" in taper_end:
                duration = round(interpolate_value(start["duration"], peak["duration"], taper_end["duration"], cycle_idx, total_cycles, peak_cycle_idx))
            if duration is None:
                duration = desc.get("duration")
            if mileage is None:
                mileage = desc.get("miles")
                if mileage is not None:
                    mileage_constrained = True
            ws.append([
                workout_id,
                microcycle_id,
                intra_idx,
                date.strftime("%Y-%m-%d"),
                weekday_formula,
                wt,
                mileage if mileage_constrained and mileage is not None else "",
                duration if not mileage_constrained and duration is not None else "",
            ])
            # Track microcycle row ranges
            if current_microcycle_start_row is None:
                current_microcycle_start_row = ws.max_row  # first row of this microcycle
            week_data.append((mileage if mileage_constrained and mileage is not None else 0,
                              duration if duration is not None else 0))
            microcycle_data.append((mileage if mileage_constrained and mileage is not None else 0,
                                   duration if duration is not None else 0))
            workout_id += 1
            if (intra_idx + 1) % days_per_microcycle == 0:
                # End of microcycle
                microcycle_row_ranges.append((current_microcycle_start_row, ws.max_row))
                current_microcycle_start_row = None
                microcycle_id += 1
                total_miles = sum(x[0] for x in microcycle_data)
                total_mins = sum(x[1] for x in microcycle_data)
                total_hours = round(total_mins / 60, 2)
                microcycle_summaries.append((microcycle_id-1, total_miles, total_hours, None, None))
                microcycle_data = []
            if date.weekday() == 6:
                total_miles = sum(x[0] for x in week_data)
                total_mins = sum(x[1] for x in week_data)
                total_hours = round(total_mins / 60, 2)
                week_summaries.append((None, total_miles, total_hours, None, None))
                week_data = []
        # If the microcycle ended early (last microcycle is short)
        if current_microcycle_start_row is not None:
            microcycle_row_ranges.append((current_microcycle_start_row, ws.max_row))
            current_microcycle_start_row = None
            total_miles = sum(x[0] for x in microcycle_data)
            total_mins = sum(x[1] for x in microcycle_data)
            total_hours = round(total_mins / 60, 2)
            microcycle_summaries.append((microcycle_id, total_miles, total_hours, None, None))
            microcycle_data = []
        if date > race_date:
            break
    # After main table, change the last 'L' to 'E' if it is within 7 days before race day
    if last_l_row and last_l_date and (0 < (race_date - last_l_date).days <= 7):
        ws[f"F{last_l_row}"] = "E"
        # Set mileage and duration to match 'E' run
        e_desc = run_descriptions.get("E", {})
        # Use interpolated or fallback duration for 'E'
        e_peak = e_desc.get("peak", {})
        e_start = e_desc.get("start", {})
        e_taper_end = e_desc.get("taper_end", {})
        # Use the same cycle index as the last long run for interpolation
        e_duration = None
        e_mileage = None
        if "duration" in e_peak and "duration" in e_start and "duration" in e_taper_end:
            e_duration = round(interpolate_value(e_start["duration"], e_peak["duration"], e_taper_end["duration"], cycle_idx, total_cycles, peak_cycle_idx))
        if "miles" in e_peak and "miles" in e_start and "miles" in e_taper_end:
            e_mileage = round(interpolate_value(e_start["miles"], e_peak["miles"], e_taper_end["miles"], cycle_idx, total_cycles, peak_cycle_idx), 2)
        if e_duration is None:
            e_duration = e_desc.get("duration", "")
        if e_mileage is None:
            e_mileage = e_desc.get("miles", "")
        ws[f"G{last_l_row}"] = e_mileage if e_mileage != "" else ""
        ws[f"H{last_l_row}"] = e_duration if e_duration != "" else ""

    # Add microcycle summaries
    ws.append([])
    ws.append(["Microcycle Summaries"])
    summary_headers = ["Microcycle ID", "Miles for Mileage-constrained", "Total Duration (hours)", "Est. Overall Mileage (formula)"]
    ws.append(summary_headers)
    for cell in ws[ws.max_row]:
        cell.alignment = Alignment(wrap_text=True)
        cell.font = Font(bold=True)
    # Use the tracked row ranges for summary formulas
    for i, (mid, _, _, _, _) in enumerate(microcycle_summaries):
        summary_row = ws.max_row + 1
        if i < len(microcycle_row_ranges):
            microcycle_start, microcycle_end = microcycle_row_ranges[i]
            if microcycle_start <= microcycle_end:
                mileage_formula = f"=SUM(G{microcycle_start}:G{microcycle_end})"
                duration_formula = f"=SUM(H{microcycle_start}:H{microcycle_end})/60"
                overall_formula = f"=B{summary_row} + (C{summary_row}*60)/$B$13"
                ws.append([mid, mileage_formula, duration_formula, overall_formula])

    ws.append([])
    ws.append(["Weekly Summaries"])
    week_headers = ["Week #", "Miles for Mileage-constrained", "Total Duration (hours)", "Est. Overall Mileage (formula)"]
    ws.append(week_headers)
    for cell in ws[ws.max_row]:
        cell.alignment = Alignment(wrap_text=True)
        cell.font = Font(bold=True)
    # For each week summary, calculate the row range in the main table
    week_start = microcycle_row_ranges[0][0] if microcycle_row_ranges else 1
    week_number = 1
    for i, (week, miles, hours, _, _) in enumerate(week_summaries):
        summary_row = ws.max_row + 1
        week_end = week_start + 6
        mileage_formula = f"=SUM(G{week_start}:G{week_end})"
        duration_formula = f"=SUM(H{week_start}:H{week_end})/60"
        overall_formula = f"=B{summary_row} + (C{summary_row}*60)/$B$13"
        ws.append([week_number, mileage_formula, duration_formula, overall_formula])
        week_start = week_end + 1
        week_number += 1

    wb.save("training_plan.xlsx")
    print("Training plan saved to training_plan.xlsx")

if __name__ == "__main__":
    main()
